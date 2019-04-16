# -*- coding: utf-8 -*-
#识别图像中的字符串 使用baidu-aip模块  优点是：快、准、简洁 
#baidu-aip模块 网站 https://console.bce.baidu.com/ai/?_=1530048735638&fromai=1#/ai/ocr/app/detail~appId=416472
#图片转文字工具 http://www.gaitubao.com/tupian-wenzi/
# 微信小程序在开发工具上执行POST请求这里(服务器)可以接受到数据；但是手机预览时执行POST请求这里(服务器)接受不到数据
# 原因：域名信息未设置 https://blog.csdn.net/ycocol/article/details/79295504

from django.shortcuts import render
from django.http.response import HttpResponseRedirect, HttpResponse,\
    StreamingHttpResponse
from django.http import JsonResponse   
from aip import AipOcr   # pip install  baidu-aip
import os,time,datetime,shutil
  
DEL_OBSOLETE = 10*60 #设置过时文件10分钟

UPFILE_SIZE = 8000000 #设置上传文件大小(8MB)
IMG_NAME = './static/img_name.jpg'
BAR_CODE_PATH = './home/static/xls'
TMP_BAR_CODE_PATH = '%s/tmp' %(BAR_CODE_PATH)

AppID = '11450107'
API_Key = 'uAA5KGXAzDDk2ewBa3dvRrWj'
Secret_Key = 'CgLyabrL4WH5KV0yyT074cMx8GyAnRGt'

#批量生成条形码
# pip install openpyxl
# pip install pystrich
from openpyxl import load_workbook,Workbook
from openpyxl.drawing.image import Image
from pystrich.code128 import Code128Encoder

#  http://localhost:8000  
def index(request):
    return  render(request, 'index.html', context=locals())    

# 保存上传的文件
def save_upfile(upfile):
    ret, fp = False, open(IMG_NAME,'wb')  
    if upfile.size < UPFILE_SIZE: #限制上传文件大小          
        for chunk in upfile.chunks(): 
            fp.write(chunk)       
        ret = True
    fp.close()
    return ret

def tostr(request):       
    if request.method == 'POST':        
        upfile = request.FILES.get("upfile", None)                              
        res = get_distinguish_img_str(IMG_NAME) if save_upfile(upfile) \
            else '上传文件大于 %s MB' %(UPFILE_SIZE/1000000)       
        shutil.copy(IMG_NAME,'./static_common')
        return  render(request, 'tostr.html', context=locals())      
    return  render(request, 'recog.html', context=locals())

#  http://localhost:8000/recog_img/  
def recog_img(request):         
    res = get_distinguish_img_str(IMG_NAME)
    mylist = [{"res" : res}] 
    return JsonResponse(mylist, safe = False) 

def apidict(request):
    mydict ={"result":[{"catid":"1","upid":"0",\
        "catname":"PhoneGap\u5e73\u53f0\u8d44\u8baf","subcate":[]},\
        {"catid":"8","upid":"0","catname":"\u89c6\u9891\u6559\u7a0b",\
         "subcate":[{"catid":"9","upid":"8",\
        "catname":"PhoneGap \u89c6\u5c4f\u6559\u7a0b"},\
        {"catid":"10","upid":"8","catname":"Sencha Touch \u6559\u7a0b"},\
        {"catid":"11","upid":"8","catname":"html5 \u89c6\u9891\u6559\u7a0b"},\
        {"catid":"12","upid":"8","catname":"jquery mobile \u6559\u7a0b"},\
        {"catid":"13","upid":"8","catname":"js\/jquery \u6559\u7a0b"},\
        {"catid":"14","upid":"8","catname":"css \u6559\u7a0b"},\
        {"catid":"26","upid":"8","catname":"JqMobi \u89c6\u9891\u6559\u7a0b"},\
        {"catid":"28","upid":"8","catname":"ionic \u6559\u7a0b"},\
        {"catid":"29","upid":"8","catname":"angularjs\u6559\u7a0b"}]},\
        {"catid":"15","upid":"0","catname":"\u79fb\u52a8\u4e92\u8054\u7f51",\
        "subcate":[]},{"catid":"20","upid":"0","catname":"PhoneGap\u8d44\u8baf",\
        "subcate":[]},{"catid":"25","upid":"0","catname":"phonegap100\u8d5e\u52a9\u89c6\u9891\u6559\u7a0b","subcate":[]},{"catid":"27","upid":"0","catname":"\u6742\u8c08","subcate":[]}]}
    return JsonResponse(mydict, safe = False) 

def apidata(request):
    res = 'hello world !'
    if request.method == 'POST':
        res += request.POST['res']
    mylist = [{"res" : res}] 
    return JsonResponse(mylist, safe = False) 

#  http://localhost:8000/distinguish_img/  
def distinguish_img(request):  
    res = get_distinguish_img_str(IMG_NAME)
    if os.path.exists(IMG_NAME):
        os.remove(IMG_NAME)
    mylist = [{"res" : res}] 
    return JsonResponse(mylist, safe = False) 

#  http://localhost:8000/wx_uploadFile/  
def wx_uploadFile(request):
    name = '' 
    if request.method == 'POST':   
        upfile = request.FILES.get("upfile", None)                              
        if upfile:
            name = myfile.name if save_upfile(upfile) \
            else '上传文件大于 %s MHZ' %(UPFILE_SIZE/1000000)                    
        else:                 
            name = 'None'         
    mylist = [{"name" : name}] 
    return JsonResponse(mylist, safe = False) 

def get_distinguish_img_str(name):
    s = ''
    try:
        client = AipOcr(AppID,API_Key,Secret_Key)
        img = open(name,'rb').read()
        msg = client.basicGeneral(img)
        for m in msg.get('words_result'):
            s += m.get('words') + '\n'
    except Exception as ex:
        s = str(ex)
    if not s:
        s = 'No Img Data !'
    return s


#条形码 http://localhost:8000/bar_code/    
def bar_code(request):
    return  render(request, 'bar_code.html', context=locals())

# 测试 怪事： http://localhost:8000/test/    2019.01.25
# 怪事：找不到对应的"bar-height"  <input id="bar-height" type="range" min="10" max="150" step="5" value="80"/>
def test(request):    
    return  render(request, 'test.html', context=locals())

# 删除目录dirname下所有的过时文件
def delallObsolete(dirname):
    file_list = os.listdir(dirname) #获取dirname下的目录和文件
    for name in file_list:
        delObsolete(name)
    return ''

# 删除过时文件。过时文件：当前时间-文件创建时间 > 设定时间
def delObsolete(name):    
    filepath = TMP_BAR_CODE_PATH + '/' + name
    cratetime = os.path.getctime(filepath)#获取文件的创建时间 1548677833.0
    nowTime = time.time()
    if nowTime > cratetime + DEL_OBSOLETE:
        os.remove(filepath)
    return ''    
    
# file电子表格文件名 批量生成条码 
def excelCode128(file,bar_width,bar_height,bar_fontSize,bar_margin):
    '''
    批量生成条码
    file(test.xlsx)   excel A列 商品编码 文件名。
    file1(pic.xlsx)   excel A列生成的条形码。
    '''
    wb = load_workbook(file)
    sheet = wb.active
    
    file1 = BAR_CODE_PATH + 'pic.xlsx'
    if os.path.exists(file1):
        os.remove(file1)
    shutil.copyfile(BAR_CODE_PATH + 'start.xlsx',file1)
    wb1 = load_workbook(file1)
    sheet1 = wb1.active
    f = []
    for r in range(1,sheet.max_row+1):
        # A列数据生成条码
        var = sheet.cell(column=1,row=r).value
        filename = str(var)+'.png'
        #ttf_fontsize:字号、label_border：行距（条码与字之间的距离）、bottom_border：字与下边框之间的距离、height：矩形框高度、bar_width:矩形框宽度
        encoder = Code128Encoder(str(var),\
            options={"ttf_fontsize":bar_fontSize,"bottom_border":5,\
                     "height":bar_height, "label_border":bar_margin})
        encoder.save(filename, bar_width=bar_width)
        #Code128Encoder(str(var)).save(filename,bar_width=1
        # 另一个表 A列数据插入条码
        img = Image(filename)
        A = 'A'+str(r)
        sheet1.add_image(img,A)
        f.append(filename)                                                                     
    wb1.save(file1)
    [os.remove(filename) for filename in f]
    print("批量生成条码成功！共计%s条！" %str(len(f)))     

 
def Code128(codestr,bar_width,bar_height,bar_fontSize,bar_margin):
    '''
    电子表格，A列存放生成的条形码
    '''    
    #NOW_TIME = datetime.datetime.now().strftime('%Y%m%d[%H_%M_%S]')  # 现在时间做文件名
    NOW_TIME = datetime.datetime.now().strftime('%H_%M_%S')
    file1 =  '%s/%s.xlsx' %(TMP_BAR_CODE_PATH,NOW_TIME)
    delallObsolete(TMP_BAR_CODE_PATH)  #删除目录TMP_BAR_CODE_PATH下的过时文件
    shutil.copyfile(BAR_CODE_PATH + '/start.xlsx',file1)
    wb1 = load_workbook(file1)
    sheet1 = wb1.active
    f = []        
    for r in codestr.split('\n'):
        r = r.strip()
        if r:
            f = en_coder(r,f,sheet1,bar_width,bar_height,bar_fontSize,bar_margin)
    wb1.save(file1)
    [os.remove(filename) for filename in f] #删除临时文件
    createNum = str(len(f))
    print("批量生成条码成功！共计%s条！" %createNum)   
    return createNum,NOW_TIME


def en_coder(r,f,sheet1,bar_width,bar_height,bar_fontSize,bar_margin):
    '''
    数据生成条码
    ttf_fontsize:字号、label_border：行距（条码与字之间的距离）、bottom_border：字与下边框之间的距离、
    height：矩形框高度、bar_width:矩形框宽度
    '''
    filename = '%s.png' %(r)
    encoder = Code128Encoder(r,options={"ttf_fontsize":bar_fontSize,\
        "bottom_border":5,"height":bar_height, "label_border":bar_margin})
    encoder.save(filename, bar_width=bar_width)

    # 电子表格 A列数据插入条码
    img = Image(filename)
    f.append(filename) 
    x = len(f)  
    A = 'A'+str(x)
    sheet1.add_image(img,A)
    return f
                                                                                 
#批量生成条形码 http://localhost:8000/batch_code/   
def batch_code(request):
    if request.method == 'POST':
        cleanData = request.POST.dict()
        userInput = cleanData.get('userInput','') #数据
        barcodeType = cleanData.get('barcodeType','') #生成条形码类型
        
        bar_width = int(cleanData.get('bar_width','0')) #宽度
        bar_height = int(cleanData.get('bar_height','0')) #高度
        bar_fontSize = int(cleanData.get('bar_fontSize','0')) #字号
        bar_margin = int(cleanData.get('bar_margin','0'))  #行距

        if barcodeType == 'Code128Encoder':
            createNum,nowtime = Code128(userInput,bar_width,bar_height,bar_fontSize,bar_margin)
        return  render(request, 'batch_code_down.html', context=locals())            
    return  render(request, 'batch_code.html', context=locals())    

def downFile(filename):
    downfilename ='%s%s' %(datetime.datetime.now().strftime('%H_%M_%S'),\
                               os.path.splitext(filename)[1] ) 
    def file_iterator(file_name, chunk_size=512):
        with open(file_name, 'rb') as f:    
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break
    response = StreamingHttpResponse(file_iterator(filename))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(downfilename)
    return response


#文件下载  http://localhost:8000/downLoadFile/
def downLoadFile(request):
    filename ='%s/%s.xlsx' %(TMP_BAR_CODE_PATH, request.GET.get('nowtime',''))
    return downFile(filename)


def QR_code(request):
    return  render(request, 'QR_code.html', context=locals())

